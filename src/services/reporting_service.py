import os
import json
import shutil
import re
from datetime import datetime
from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, PageBreak
)
from src.services.db_service import (
    update_pipeline_stage, log_activity, Dataset, PipelineStage, db
)
from src.services.visualization_service import _sanitize_name, _chart_dir
from src.utils.logger import pipeline_logger, error_logger

HEADER_FILL = "1F2937"
ACCENT_HEX = "3B82F6"


def _sanitize_folder(name: str) -> str:
    return _sanitize_name(name)


def _report_dir(dataset_name: str, dataset_uuid: str) -> str:
    base = current_app.config['REPORTS_DIR']
    year = str(datetime.utcnow().year)
    folder = f"{_sanitize_folder(dataset_name)}_{dataset_uuid[:8]}"
    report_dir = os.path.join(base, year, folder)
    os.makedirs(report_dir, exist_ok=True)
    return report_dir


def _load_json(path):
    if path and os.path.exists(path):
        with open(path, 'r') as f:
            return json.load(f)
    return None


def _gather_context(dataset_uuid: str) -> dict:
    """Collects all pipeline artifacts (validation, cleaning, analytics, charts) for a dataset."""
    dataset = Dataset.query.filter_by(uuid=dataset_uuid).first()
    if not dataset:
        raise ValueError(f"Dataset with UUID {dataset_uuid} not found.")

    validation_summary = None
    if dataset.raw_filepath:
        validation_summary = _load_json(os.path.join(os.path.dirname(dataset.raw_filepath), 'validation_summary.json'))

    cleaning_summary = None
    if dataset.cleaned_filepath:
        cleaning_summary = _load_json(os.path.join(os.path.dirname(dataset.cleaned_filepath), 'cleaning_summary.json'))

    analytics_summary = _load_json(
        os.path.join(current_app.config['BASE_DIR'], 'uploads', 'analytics', dataset_uuid, 'analytics_summary.json')
    )

    chart_dir = _chart_dir(dataset.dataset_name, dataset_uuid)
    chart_manifest = _load_json(os.path.join(chart_dir, 'manifest.json'))

    stages = PipelineStage.query.filter_by(dataset_id=dataset.id).all()

    return {
        'dataset': dataset,
        'validation_summary': validation_summary,
        'cleaning_summary': cleaning_summary,
        'analytics_summary': analytics_summary,
        'chart_manifest': chart_manifest,
        'chart_dir': chart_dir,
        'stages': stages,
    }


# ---------------------------------------------------------------------------
# Excel Report
# ---------------------------------------------------------------------------

def _style_header_row(ws, row_idx, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize_columns(ws, max_width=60):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(12, length + 2), max_width)


def generate_excel_report(dataset_uuid: str) -> dict:
    """Generates a multi-sheet Excel workbook summarizing the full pipeline run for a dataset."""
    update_pipeline_stage(dataset_uuid, 'Excel_Report', 'Running')

    try:
        ctx = _gather_context(dataset_uuid)
        dataset = ctx['dataset']

        wb = Workbook()

        # --- Sheet 1: Dataset Info ---
        ws_info = wb.active
        ws_info.title = "Dataset Info"
        ws_info.append(["Field", "Value"])
        _style_header_row(ws_info, 1, 2)
        info_rows = [
            ("Dataset Name", dataset.dataset_name),
            ("Original Filename", dataset.original_filename),
            ("UUID", dataset.uuid),
            ("Row Count", dataset.row_count),
            ("Column Count", dataset.column_count),
            ("File Size (bytes)", dataset.file_size_bytes),
            ("Upload Time", dataset.upload_time.strftime('%Y-%m-%d %H:%M:%S') if dataset.upload_time else ""),
            ("Overall Status", dataset.status),
            ("Report Generated", datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')),
        ]
        for row in info_rows:
            ws_info.append(row)
        _autosize_columns(ws_info)

        # --- Sheet 2: Validation Summary ---
        ws_val = wb.create_sheet("Validation Summary")
        vs = ctx['validation_summary'] or {}
        ws_val.append(["Metric", "Value"])
        _style_header_row(ws_val, 1, 2)
        ws_val.append(["Status", vs.get('status', 'N/A')])
        ws_val.append(["Errors", len(vs.get('errors', []))])
        ws_val.append(["Warnings", len(vs.get('warnings', []))])
        ws_val.append(["Missing Required Columns", ', '.join(vs.get('missing_required_columns', [])) or 'None'])
        ws_val.append([])
        ws_val.append(["Warning Details"])
        for w in vs.get('warnings', []):
            ws_val.append([w])
        for e in vs.get('errors', []):
            ws_val.append([f"ERROR: {e}"])
        _autosize_columns(ws_val)

        # --- Sheet 3: Cleaning Summary ---
        ws_clean = wb.create_sheet("Cleaning Summary")
        cs = ctx['cleaning_summary'] or {}
        ws_clean.append(["Metric", "Value"])
        _style_header_row(ws_clean, 1, 2)
        for label, key in [
            ("Original Rows", 'original_rows'), ("Cleaned Rows", 'cleaned_rows'),
            ("Empty Rows Removed", 'empty_rows_removed'), ("Duplicates Removed", 'duplicates_removed'),
            ("Nulls Filled", 'nulls_filled'), ("Casing Rule", 'casing_rule'),
        ]:
            ws_clean.append([label, cs.get(key, 'N/A')])
        _autosize_columns(ws_clean)

        # --- Sheet 4: Analytics KPIs ---
        ws_kpi = wb.create_sheet("Analytics KPIs")
        an = ctx['analytics_summary'] or {}
        ws_kpi.append(["KPI", "Value"])
        _style_header_row(ws_kpi, 1, 2)
        ws_kpi.append(["Row Count", an.get('row_count', 'N/A')])
        ws_kpi.append(["Column Count", an.get('column_count', 'N/A')])
        ws_kpi.append(["Missing %", an.get('missing_percentage', 'N/A')])
        rev = an.get('revenue_metrics', {})
        if rev.get('available'):
            ws_kpi.append(["Total Revenue", rev.get('total_revenue')])
            ws_kpi.append(["Average Revenue", rev.get('average_revenue')])
            ws_kpi.append(["Max Revenue", rev.get('max_revenue')])
            ws_kpi.append(["Min Revenue", rev.get('min_revenue')])
        sales = an.get('sales_metrics', {})
        if sales.get('available'):
            ws_kpi.append(["Total Units Sold", sales.get('total_units_sold')])
            ws_kpi.append(["Average Units Sold", sales.get('average_units_sold')])
        _autosize_columns(ws_kpi)

        # --- Sheet 5: Top Categories ---
        ws_cat = wb.create_sheet("Top Categories")
        ws_cat.append(["Category", "Value", "Metric", "Percentage"])
        _style_header_row(ws_cat, 1, 4)
        for cat in an.get('top_categories', []):
            ws_cat.append([cat.get('category'), cat.get('value'), cat.get('metric'), cat.get('percentage')])
        _autosize_columns(ws_cat)

        # --- Sheet 6: Monthly Trends ---
        ws_trend = wb.create_sheet("Monthly Trends")
        ws_trend.append(["Month", "Value", "Metric"])
        _style_header_row(ws_trend, 1, 3)
        for point in an.get('monthly_trends', []):
            ws_trend.append([point.get('month'), point.get('value'), point.get('metric')])
        _autosize_columns(ws_trend)

        # --- Sheet 7: Summary Statistics ---
        ws_stats = wb.create_sheet("Summary Statistics")
        ws_stats.append(["Column", "Sum", "Mean", "Median", "Std Dev", "Min", "Max", "Count"])
        _style_header_row(ws_stats, 1, 8)
        for col, stats in an.get('summary_statistics', {}).items():
            ws_stats.append([col, stats.get('sum'), stats.get('mean'), stats.get('median'),
                              stats.get('std_dev'), stats.get('min'), stats.get('max'), stats.get('count')])
        _autosize_columns(ws_stats)

        # --- Sheet 8: Pipeline Summary ---
        ws_pipe = wb.create_sheet("Pipeline Summary")
        ws_pipe.append(["Stage", "Status", "Started At", "Completed At", "Error"])
        _style_header_row(ws_pipe, 1, 5)
        for stage in ctx['stages']:
            ws_pipe.append([
                stage.stage_name, stage.status,
                stage.started_at.strftime('%Y-%m-%d %H:%M:%S') if stage.started_at else '',
                stage.completed_at.strftime('%Y-%m-%d %H:%M:%S') if stage.completed_at else '',
                stage.error_message or ''
            ])
        _autosize_columns(ws_pipe)

        # --- Sheet 9: Charts (embedded PNGs) ---
        manifest = ctx['chart_manifest'] or {}
        charts = manifest.get('charts', [])
        if charts:
            ws_charts = wb.create_sheet("Charts")
            row_cursor = 1
            for chart in charts:
                png_path = os.path.join(ctx['chart_dir'], chart['png'])
                ws_charts.cell(row=row_cursor, column=1, value=chart['title']).font = Font(bold=True, color=ACCENT_HEX)
                row_cursor += 1
                if os.path.exists(png_path):
                    img = XLImage(png_path)
                    img.width = 480
                    img.height = 300
                    ws_charts.add_image(img, f"A{row_cursor}")
                row_cursor += 18

        report_dir = _report_dir(dataset.dataset_name, dataset_uuid)
        excel_path = os.path.join(report_dir, 'report.xlsx')
        wb.save(excel_path)

        dataset_id = dataset.id
        update_pipeline_stage(dataset_uuid, 'Excel_Report', 'Completed')

        log_msg = f"Excel report generated with {len(wb.sheetnames)} sheet(s) at '{excel_path}'."
        pipeline_logger.info(log_msg)
        log_activity(log_msg, level='INFO', dataset_id=dataset_id, stage='Excel_Report')

        return {'excel_path': excel_path, 'report_dir': report_dir}
    except Exception as e:
        error_msg = f"Fatal Excel report generation failure: {str(e)}"
        error_logger.error(error_msg)
        update_pipeline_stage(dataset_uuid, 'Excel_Report', 'Failed', error_message=error_msg)
        raise e


# ---------------------------------------------------------------------------
# PDF Report
# ---------------------------------------------------------------------------

def generate_pdf_report(dataset_uuid: str) -> dict:
    """Generates a professional PDF report summarizing the full pipeline run for a dataset."""
    update_pipeline_stage(dataset_uuid, 'PDF_Report', 'Running')

    try:
        ctx = _gather_context(dataset_uuid)
        dataset = ctx['dataset']
        vs = ctx['validation_summary'] or {}
        cs = ctx['cleaning_summary'] or {}
        an = ctx['analytics_summary'] or {}
        manifest = ctx['chart_manifest'] or {}

        report_dir = _report_dir(dataset.dataset_name, dataset_uuid)
        pdf_path = os.path.join(report_dir, 'report.pdf')

        doc = SimpleDocTemplate(pdf_path, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('TitleCustom', parent=styles['Title'], textColor=colors.HexColor("#111827"))
        heading_style = ParagraphStyle('HeadingCustom', parent=styles['Heading2'], textColor=colors.HexColor("#1F2937"),
                                        spaceBefore=14, spaceAfter=6)
        body_style = styles['BodyText']

        elements = []
        elements.append(Paragraph("DA Automation System (DAAS) - Pipeline Report", title_style))
        elements.append(Paragraph(f"Dataset: {dataset.dataset_name}", styles['Heading3']))
        elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC", body_style))
        elements.append(Spacer(1, 12))

        # --- Dataset Information ---
        elements.append(Paragraph("1. Dataset Information", heading_style))
        info_table_data = [
            ["Original Filename", dataset.original_filename],
            ["UUID", dataset.uuid],
            ["Row Count", str(dataset.row_count)],
            ["Column Count", str(dataset.column_count)],
            ["File Size", f"{(dataset.file_size_bytes or 0) / 1024:.2f} KB"],
            ["Overall Status", dataset.status],
        ]
        elements.append(_styled_table(info_table_data))

        # --- Validation Summary ---
        elements.append(Paragraph("2. Validation Summary", heading_style))
        val_data = [
            ["Status", vs.get('status', 'N/A')],
            ["Errors", str(len(vs.get('errors', [])))],
            ["Warnings", str(len(vs.get('warnings', [])))],
            ["Missing Required Columns", ', '.join(vs.get('missing_required_columns', [])) or 'None'],
        ]
        elements.append(_styled_table(val_data))

        # --- Cleaning Summary ---
        elements.append(Paragraph("3. Cleaning Summary", heading_style))
        clean_data = [
            ["Original Rows", str(cs.get('original_rows', 'N/A'))],
            ["Cleaned Rows", str(cs.get('cleaned_rows', 'N/A'))],
            ["Empty Rows Removed", str(cs.get('empty_rows_removed', 'N/A'))],
            ["Duplicates Removed", str(cs.get('duplicates_removed', 'N/A'))],
            ["Nulls Filled", str(cs.get('nulls_filled', 'N/A'))],
        ]
        elements.append(_styled_table(clean_data))

        # --- KPIs ---
        elements.append(Paragraph("4. Analytics KPIs", heading_style))
        kpi_data = [["Row Count", str(an.get('row_count', 'N/A'))], ["Missing %", f"{an.get('missing_percentage', 'N/A')}%"]]
        rev = an.get('revenue_metrics', {})
        if rev.get('available'):
            kpi_data += [
                ["Total Revenue", f"${rev.get('total_revenue', 0):,.2f}"],
                ["Average Revenue", f"${rev.get('average_revenue', 0):,.2f}"],
            ]
        sales = an.get('sales_metrics', {})
        if sales.get('available'):
            kpi_data += [["Total Units Sold", f"{sales.get('total_units_sold', 0):,.0f}"]]
        elements.append(_styled_table(kpi_data))

        # --- Charts ---
        charts = manifest.get('charts', [])
        if charts:
            elements.append(PageBreak())
            elements.append(Paragraph("5. Visualizations", heading_style))
            for chart in charts:
                png_path = os.path.join(ctx['chart_dir'], chart['png'])
                if os.path.exists(png_path):
                    elements.append(Paragraph(chart['title'], styles['Heading4']))
                    elements.append(RLImage(png_path, width=15 * cm, height=9 * cm))
                    elements.append(Spacer(1, 10))

        # --- Pipeline Summary ---
        elements.append(PageBreak())
        elements.append(Paragraph("6. Pipeline Summary", heading_style))
        pipe_data = [["Stage", "Status", "Completed At"]]
        for stage in ctx['stages']:
            pipe_data.append([
                stage.stage_name, stage.status,
                stage.completed_at.strftime('%Y-%m-%d %H:%M:%S') if stage.completed_at else '-'
            ])
        elements.append(_styled_table(pipe_data, header=True))

        doc.build(elements)

        dataset_id = dataset.id
        update_pipeline_stage(dataset_uuid, 'PDF_Report', 'Completed')

        log_msg = f"PDF report generated at '{pdf_path}'."
        pipeline_logger.info(log_msg)
        log_activity(log_msg, level='INFO', dataset_id=dataset_id, stage='PDF_Report')

        return {'pdf_path': pdf_path, 'report_dir': report_dir}
    except Exception as e:
        error_msg = f"Fatal PDF report generation failure: {str(e)}"
        error_logger.error(error_msg)
        update_pipeline_stage(dataset_uuid, 'PDF_Report', 'Failed', error_message=error_msg)
        raise e


def _styled_table(data, header=False):
    table = Table(data, colWidths=[6 * cm, 9 * cm] if len(data[0]) == 2 else None, hAlign='LEFT')
    style_cmds = [
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]
    if header:
        style_cmds += [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F2937")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]
    else:
        style_cmds += [('BACKGROUND', (0, 0), (0, -1), colors.HexColor("#f3f4f6")),
                        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold')]
    table.setStyle(TableStyle(style_cmds))
    return table


# ---------------------------------------------------------------------------
# Archiving (final pipeline stage)
# ---------------------------------------------------------------------------

def archive_dataset(dataset_uuid: str) -> dict:
    """Copies the raw dataset file into the archive directory, marking the pipeline as complete."""
    update_pipeline_stage(dataset_uuid, 'Archive', 'Running')

    try:
        dataset = Dataset.query.filter_by(uuid=dataset_uuid).first()
        if not dataset:
            raise ValueError(f"Dataset with UUID {dataset_uuid} not found.")

        archive_dir = os.path.join(current_app.config['UPLOAD_ARCHIVE_DIR'], dataset_uuid)
        os.makedirs(archive_dir, exist_ok=True)

        archive_path = os.path.join(archive_dir, dataset.original_filename)
        if dataset.raw_filepath and os.path.exists(dataset.raw_filepath) and not os.path.exists(archive_path):
            shutil.copy2(dataset.raw_filepath, archive_path)

        dataset.archive_filepath = archive_path
        db.session.commit()

        update_pipeline_stage(dataset_uuid, 'Archive', 'Completed')

        log_msg = f"Dataset archived to '{archive_path}'. Pipeline run complete."
        pipeline_logger.info(log_msg)
        log_activity(log_msg, level='INFO', dataset_id=dataset.id, stage='Archive')

        return {'archive_path': archive_path}
    except Exception as e:
        error_msg = f"Fatal archiving failure: {str(e)}"
        error_logger.error(error_msg)
        update_pipeline_stage(dataset_uuid, 'Archive', 'Failed', error_message=error_msg)
        raise e
